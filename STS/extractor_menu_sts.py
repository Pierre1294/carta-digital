#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor de Menú - Oracle Simphony STS (Gen2)
Consulta el servicio STS getMenuItems y exporta a Excel
Autor: Script para Pierre - Popeyes Perú
Versión: 1.0

NOTA IMPORTANTE:
Este endpoint aún no ha sido validado con evidencia de Postman, por lo que
la extracción de campos se hace de forma GENÉRICA (recursiva/aplanada) para
no depender de una estructura JSON asumida. Además se guarda el JSON crudo
de la respuesta para poder ajustar el mapeo de campos si hace falta.
"""

import json
import logging
import time
from datetime import datetime
from pathlib import Path

import requests
import openpyxl
from openpyxl.utils import get_column_letter


class ExtractorMenuSTS:
    def __init__(self, config_path='config_extractor_menu_sts.json'):
        self.config = self._cargar_config(config_path)
        self._configurar_logging()
        logging.info("=" * 70)
        logging.info("EXTRACTOR DE MENÚ - ORACLE SIMPHONY STS (Gen2)")
        logging.info("=" * 70)

    # ------------------------------------------------------------------ #
    # Configuración / logging
    # ------------------------------------------------------------------ #
    def _cargar_config(self, config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"ERROR: No se encontró el archivo de configuración: {config_path}")
            raise
        except json.JSONDecodeError as e:
            print(f"ERROR: JSON inválido en configuración: {e}")
            raise

    def _configurar_logging(self):
        log_config = self.config['logging']
        logging.basicConfig(
            level=getattr(logging, log_config['nivel']),
            format=log_config['formato'],
            handlers=[
                logging.FileHandler(log_config['archivo_log'], encoding='utf-8'),
                logging.StreamHandler()
            ]
        )

    # ------------------------------------------------------------------ #
    # Construcción de URL y llamada al API
    # ------------------------------------------------------------------ #
    def _construir_menu_id(self):
        org = self.config['api']['org']
        env_loc_ref = self.config['menuId']['env_locRef']
        loc_ref = self.config['menuId']['locRef']
        return f"{org}:{env_loc_ref}:{loc_ref}"

    def _construir_url(self, menu_id):
        base_url = self.config['api']['base_url']
        endpoint = self.config['api']['endpoint'].format(menuId=menu_id)
        return f"{base_url}{endpoint}"

    def _obtener_headers(self):
        headers = {'Content-Type': 'application/json'}
        token = self.config['api'].get('auth', {}).get('token', '').strip()
        if token and token != 'TU_TOKEN_AQUI':
            headers['Authorization'] = f'Bearer {token}'
        else:
            logging.warning("  ⚠️  No hay token configurado en 'api.auth.token'. "
                             "Si el endpoint responde 401, configúralo.")

        env_loc_ref = self.config['menuId']['env_locRef']
        loc_ref = self.config['menuId']['locRef']

        headers_extra = self.config['api'].get('headers_adicionales', {})
        for clave, valor in headers_extra.items():
            if clave == 'comentario' or not valor:
                continue
            valor_final = str(valor).replace('{env_locRef}', str(env_loc_ref)) \
                                     .replace('{locRef}', str(loc_ref))
            headers[clave] = valor_final

        return headers

    def _llamar_api(self, url, intento=1):
        try:
            headers = self._obtener_headers()
            headers_log = {k: (v if k != 'Authorization' else 'Bearer ***') for k, v in headers.items()}
            logging.info(f"  → GET {url}")
            logging.info(f"  → Headers: {headers_log}")
            response = requests.get(
                url,
                headers=headers,
                timeout=self.config['api']['timeout']
            )

            logging.info(f"  ← HTTP {response.status_code}")

            if response.status_code == 200:
                return response.json()
            elif response.status_code == 401:
                logging.error("  ❌ 401 Unauthorized: revisa que 'api.auth.token' esté "
                               "configurado correctamente en el JSON y que no haya expirado.")
                return None
            elif response.status_code == 404:
                logging.error("  ❌ 404 Not Found: revisa que env_locRef y locRef "
                               "estén bien configurados en el menuId.")
                return None
            else:
                logging.error(f"  ❌ Error HTTP {response.status_code}: {response.text[:500]}")
                return None

        except requests.exceptions.Timeout:
            if intento < self.config['api']['max_retries']:
                logging.warning(f"  ⏱️  Timeout, reintentando ({intento}/{self.config['api']['max_retries']})...")
                time.sleep(2 * intento)
                return self._llamar_api(url, intento + 1)
            logging.error("  ❌ Timeout después de varios intentos")
            return None
        except Exception as e:
            logging.error(f"  ❌ Error en la petición: {str(e)}")
            return None

    # ------------------------------------------------------------------ #
    # Guardar JSON crudo (para inspección / ajuste de mapeo)
    # ------------------------------------------------------------------ #
    def _guardar_json_crudo(self, data):
        if not self.config['output'].get('guardar_json_crudo', True):
            return

        directorio = Path(self.config['output']['directorio'])
        directorio.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        ruta = directorio / f"raw_response_{timestamp}.json"

        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        logging.info(f"  💾 JSON crudo guardado en: {ruta}")

    # ------------------------------------------------------------------ #
    # Catálogos de referencia (para resolver nombres de los *Ref)
    # ------------------------------------------------------------------ #
    # ------------------------------------------------------------------ #
    # Hoja "Catalogo" - VISTA ÚNICA JERÁRQUICA (4 niveles)
    #
    #   Nivel 1: COMBO (cabecera) ó PRODUCTO suelto
    #   Nivel 2:   └─ GRUPO del combo ("pregunta"; * = grupo principal)
    #   Nivel 3:       └─ ITEM componente del grupo ("respuesta")
    #   Nivel 4:           └─ COND (condimento; "sub-pregunta" = el grupo
    #                          de condimentos al que pertenece, con su
    #                          Min/Max heredado de esa regla)
    #
    # Solo se usa priceSequence=1 para 'Valor' (se verificó que el 100% de
    # los items con precio tienen esa secuencia, así que no se pierde nada
    # al ignorar las demás secuencias/niveles).
    # ------------------------------------------------------------------ #
    def _procesar_catalogo_jerarquico(self, data):
        idioma_es = self.config['idiomas']['nombre_espanol']
        idioma_en = self.config['idiomas']['nombre_ingles']

        menu_items = {str(i.get('menuItemId')): i for i in data.get('menuItems', [])}
        cond_groups = {str(g.get('condimentGroupId')): g for g in data.get('condimentGroups', [])}
        cond_items = {str(c.get('condimentId')): c for c in data.get('condimentItems', [])}

        def _nom_es(obj):
            """Nombre en español; si el objeto no trae es-MX (ej. los
            condimentGroups solo traen en-US), usa inglés como respaldo."""
            nombres = (obj or {}).get('name') or {}
            return nombres.get(idioma_es) or nombres.get(idioma_en, '')

        def _indentar(nivel, texto):
            return ('  ' * (nivel - 1)) + texto

        def _precio_seq1(precios):
            for p in precios or []:
                if p.get('priceSequence') == 1:
                    return p.get('price', '')
            return ''

        def _slu_y_extensiones(item):
            definiciones = item.get('definitions', []) if item else []
            if not definiciones:
                return '', ''
            d0 = definiciones[0]
            slus = '; '.join(str((s.get('name') or {}).get(idioma_en, s.get('sluId', '')))
                             for s in d0.get('slus', []))
            ext = '; '.join(f"{k}={v}" for k, v in (d0.get('extensions') or {}).items())
            return slus, ext

        def _fila_base():
            return {
                'Nivel': '', 'Tipo': '', 'Codigo': '', 'Nombre': '',
                'Valor': '', 'Principal': '', 'Min': '', 'Max': '',
                'Combo_Codigo': '', 'Combo_Nombre': '',
                'Grupo_Codigo': '', 'Grupo_Nombre': '',
                'Item_Codigo': '', 'Item_Nombre': '',
                'SubPregunta_Codigo': '', 'SubPregunta_Nombre': '',
                'SLU': '', 'Extensiones': '',
            }

        filas = []

        def _agregar_condimentos(item, nivel_item, ctx):
            """Nivel 4: condimentos, agrupados por su 'sub-pregunta'
            (el condimentGroup al que pertenecen)."""
            if not item:
                return
            for definicion in item.get('definitions', []):
                for regla in definicion.get('condimentGroupRules', []):
                    ref_grupo = str(regla.get('condimentGroupRef', ''))
                    grupo_cond = cond_groups.get(ref_grupo)
                    nombre_subpregunta = _nom_es(grupo_cond)

                    for cref in (grupo_cond.get('condimentItemRefs', []) if grupo_cond else []):
                        cond = cond_items.get(str(cref))
                        slu_c, ext_c = _slu_y_extensiones(cond)

                        fila = _fila_base()
                        fila.update(ctx)
                        fila.update({
                            'Nivel': nivel_item + 1, 'Tipo': 'COND',
                            'Codigo': cref,
                            'Nombre': _indentar(nivel_item + 1, _nom_es(cond)),
                            'Valor': _precio_seq1((cond or {}).get('definitions', [{}])[0].get('prices', []) if cond else []),
                            'Min': regla.get('minimumCount', ''), 'Max': regla.get('maximumCount', ''),
                            'SubPregunta_Codigo': ref_grupo, 'SubPregunta_Nombre': nombre_subpregunta,
                            'SLU': slu_c, 'Extensiones': ext_c,
                        })
                        filas.append(fila)

        # ---------------- Nivel 1: COMBOS ---------------- #
        combos = data.get('comboMeals', [])
        for combo in combos:
            ref_cab = str(combo.get('menuItemRef', ''))
            cabecera = menu_items.get(ref_cab)
            slu_cab, ext_cab = _slu_y_extensiones(cabecera)
            definiciones_cab = (cabecera or {}).get('definitions', [])
            precio_cab = _precio_seq1(definiciones_cab[0].get('prices', []) if definiciones_cab else [])
            nombre_cab = _nom_es(cabecera) or combo.get('name', '')

            fila1 = _fila_base()
            fila1.update({
                'Nivel': 1, 'Tipo': 'COMBO',
                'Codigo': combo.get('comboMealId', ''),
                'Nombre': nombre_cab,
                'Valor': precio_cab,
                'Combo_Codigo': combo.get('comboMealId', ''), 'Combo_Nombre': nombre_cab,
                'SLU': slu_cab, 'Extensiones': ext_cab,
            })
            filas.append(fila1)

            ctx_combo = {'Combo_Codigo': fila1['Combo_Codigo'], 'Combo_Nombre': fila1['Combo_Nombre']}

            # ---------------- Nivel 2: GRUPOS DEL COMBO ("pregunta") ---------------- #
            for grupo in combo.get('comboGroups', []):
                principal = bool(grupo.get('isMainGroup'))
                nombre_grupo = grupo.get('name', '')
                nombre_grupo_mostrado = ('* ' if principal else '') + nombre_grupo

                fila2 = _fila_base()
                fila2.update(ctx_combo)
                fila2.update({
                    'Nivel': 2, 'Tipo': 'GRUPO',
                    'Codigo': grupo.get('comboGroupId', ''),
                    'Nombre': _indentar(2, nombre_grupo_mostrado),
                    'Principal': 'Sí' if principal else 'No',
                    'Grupo_Codigo': grupo.get('comboGroupId', ''), 'Grupo_Nombre': nombre_grupo,
                })
                filas.append(fila2)

                ctx_grupo = dict(ctx_combo)
                ctx_grupo.update({'Grupo_Codigo': fila2['Grupo_Codigo'], 'Grupo_Nombre': fila2['Grupo_Nombre']})

                # ---------------- Nivel 3: ITEMS DEL GRUPO ("respuesta") ---------------- #
                for mi in grupo.get('menuItems', []):
                    ref = str(mi.get('menuItemRef', ''))
                    item = menu_items.get(ref)
                    nombre_item = _nom_es(item)
                    precio_item = _precio_seq1(mi.get('prices', []))

                    fila3 = _fila_base()
                    fila3.update(ctx_grupo)
                    fila3.update({
                        'Nivel': 3, 'Tipo': 'ITEM',
                        'Codigo': ref,
                        'Nombre': _indentar(3, nombre_item),
                        'Valor': precio_item,
                        'Item_Codigo': ref, 'Item_Nombre': nombre_item,
                    })
                    filas.append(fila3)

                    ctx_item = dict(ctx_grupo)
                    ctx_item.update({'Item_Codigo': fila3['Item_Codigo'], 'Item_Nombre': fila3['Item_Nombre']})

                    _agregar_condimentos(item, 3, ctx_item)

        # ---------------- Nivel 1: PRODUCTOS SUELTOS ---------------- #
        refs_combo = {str(c.get('menuItemRef', '')) for c in combos}
        sueltos = 0
        for codigo, item in menu_items.items():
            if codigo in refs_combo:
                continue  # ya salió como cabecera de combo
            sueltos += 1
            slu_i, ext_i = _slu_y_extensiones(item)
            definiciones_i = item.get('definitions', [])
            precio_i = _precio_seq1(definiciones_i[0].get('prices', []) if definiciones_i else [])
            nombre_i = _nom_es(item)

            fila1 = _fila_base()
            fila1.update({
                'Nivel': 1, 'Tipo': 'PRODUCTO',
                'Codigo': codigo,
                'Nombre': nombre_i,
                'Valor': precio_i,
                'Item_Codigo': codigo, 'Item_Nombre': nombre_i,
                'SLU': slu_i, 'Extensiones': ext_i,
            })
            filas.append(fila1)

            ctx1 = {'Item_Codigo': codigo, 'Item_Nombre': nombre_i}
            _agregar_condimentos(item, 1, ctx1)

        logging.info(f"  📦 {len(filas)} filas en la vista jerárquica "
                      f"({len(combos)} combos + {sueltos} productos sueltos)")

        columnas = ['Nivel', 'Tipo', 'Codigo', 'Nombre', 'Valor', 'Principal', 'Min', 'Max',
                    'Combo_Codigo', 'Combo_Nombre', 'Grupo_Codigo', 'Grupo_Nombre',
                    'Item_Codigo', 'Item_Nombre', 'SubPregunta_Codigo', 'SubPregunta_Nombre',
                    'SLU', 'Extensiones']
        return filas, columnas

    # ------------------------------------------------------------------ #
    # Generar Excel (con soporte de agrupación/colapso por nivel)
    # ------------------------------------------------------------------ #
    def _escribir_hoja(self, wb, nombre_hoja, filas, columnas, columnas_principales, columna_nivel=None):
        if not filas:
            logging.warning(f"  ⚠️  Sin datos para la hoja '{nombre_hoja}', se omite")
            return

        columnas_restantes = sorted(c for c in columnas if c not in columnas_principales)
        columnas_finales = columnas_principales + columnas_restantes

        ws = wb.create_sheet(nombre_hoja)

        for col_idx, nombre_col in enumerate(columnas_finales, start=1):
            ws.cell(1, col_idx, nombre_col)
            ws.cell(1, col_idx).font = openpyxl.styles.Font(bold=True)

        idx_nivel = columnas_finales.index(columna_nivel) if columna_nivel in columnas_finales else None

        for row_idx, fila in enumerate(filas, start=2):
            for col_idx, nombre_col in enumerate(columnas_finales, start=1):
                valor = fila.get(nombre_col, '')
                ws.cell(row_idx, col_idx, valor)

            if idx_nivel is not None:
                nivel = fila.get(columna_nivel)
                if isinstance(nivel, int) and nivel > 1:
                    ws.row_dimensions[row_idx].outlineLevel = min(nivel - 1, 7)

        for col_idx, nombre_col in enumerate(columnas_finales, start=1):
            letra = get_column_letter(col_idx)
            ws.column_dimensions[letra].width = min(max(len(nombre_col) + 2, 12), 40)

        if idx_nivel is not None:
            ws.sheet_properties.outlinePr.summaryBelow = False
            ws.sheet_properties.outlinePr.summaryRight = False

        ws.freeze_panes = "A2"
        logging.info(f"  ✓ Hoja '{nombre_hoja}' creada: {len(filas)} filas, {len(columnas_finales)} columnas")

    def _generar_excel(self, hojas):
        """
        hojas: lista de tuplas (nombre_hoja, filas, columnas, columnas_principales)
               o (nombre_hoja, filas, columnas, columnas_principales, columna_nivel)
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # quitar hoja por defecto

        for hoja in hojas:
            if len(hoja) == 5:
                nombre_hoja, filas, columnas, columnas_principales, columna_nivel = hoja
            else:
                nombre_hoja, filas, columnas, columnas_principales = hoja
                columna_nivel = None
            self._escribir_hoja(wb, nombre_hoja, filas, columnas, columnas_principales, columna_nivel)

        if not wb.sheetnames:
            logging.warning("  ⚠️  No se generó ninguna hoja, no se guarda archivo")
            return None

        directorio = Path(self.config['output']['directorio'])
        directorio.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"{self.config['output']['prefijo_archivo']}_{timestamp}.xlsx"
        ruta = directorio / nombre_archivo

        wb.save(ruta)
        logging.info(f"\n  ✅ Excel generado: {ruta}")
        return ruta

    # ------------------------------------------------------------------ #
    # Orquestador principal
    # ------------------------------------------------------------------ #
    def ejecutar(self):
        menu_id = self._construir_menu_id()
        logging.info(f"\n📋 menuId construido: {menu_id}")

        url = self._construir_url(menu_id)

        logging.info("\n🔄 Consultando STS...")
        data = self._llamar_api(url)

        if data is None:
            logging.error("\n❌ No se obtuvo respuesta válida del API. Revisa el log arriba.")
            return None

        self._guardar_json_crudo(data)

        logging.info("\n🔍 Procesando 'Catalogo' (vista jerárquica única)...")
        filas_cat, cols_cat = self._procesar_catalogo_jerarquico(data)

        if not filas_cat:
            logging.warning("\n⚠️  No se encontraron datos para armar el catálogo.")
            logging.warning("    Revisa el JSON crudo guardado para ver la estructura real.")
            return None

        hojas = [
            ("Catalogo", filas_cat, cols_cat, cols_cat, 'Nivel'),
        ]

        ruta = self._generar_excel(hojas)

        logging.info("\n" + "=" * 70)
        logging.info("✅ PROCESO COMPLETADO")
        if ruta:
            logging.info(f"   Archivo: {ruta}")
        logging.info("=" * 70)

        return ruta


def main():
    print("\n" + "=" * 70)
    print("EXTRACTOR DE MENÚ - ORACLE SIMPHONY STS (Gen2)")
    print("=" * 70 + "\n")

    try:
        extractor = ExtractorMenuSTS('config_extractor_menu_sts.json')
        extractor.ejecutar()
    except KeyboardInterrupt:
        print("\n\n⚠️  Proceso interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        print("Revisa extractor_menu_sts.log para más detalles\n")


if __name__ == "__main__":
    main()
